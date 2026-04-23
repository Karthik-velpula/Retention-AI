from __future__ import annotations

import csv
import re
import sys
from pathlib import Path

import openpyxl
from email_validator import EmailNotValidError, validate_email

from app.db.session import SessionLocal
from app.models.financial import Financial
from app.models.lms_activity import LMSActivity
from app.models.student import Student
from app.models.subject_attendance import SubjectAttendance

SOURCE_PATH = Path("/Users/karthi/Downloads/3 YEAR 1-10 SECTION ATT.xlsx")
CSV_SOURCE_PATH = Path("/Users/karthi/Downloads/studentdetails1.csv")


def _normalize_email(raw_email: str, registration_number: str) -> str:
    value = str(raw_email).strip().lower().replace(" ", "")
    if not value or value in {"-", "0", "nan"}:
        return f"{registration_number.lower()}@gmail.com"
    try:
        return validate_email(value, check_deliverability=False).normalized
    except EmailNotValidError:
        return f"{registration_number.lower()}@gmail.com"


def _csv_lookup(source_path: Path) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    with source_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            registration_number = (row.get("registerno") or "").strip()
            if registration_number:
                lookup[registration_number] = row
    return lookup


def _header_rows(ws: openpyxl.worksheet.worksheet.Worksheet) -> tuple[int, int, int]:
    header_row = None
    for row_idx in range(1, min(ws.max_row, 10) + 1):
        if ws.cell(row=row_idx, column=1).value == "SL":
            header_row = row_idx
            break
    if header_row is None:
        raise ValueError(f"Could not find header row in sheet {ws.title}.")
    code_row = header_row + 1
    conduct_row = header_row + 2
    return header_row, code_row, conduct_row


def _subject_columns(
    ws: openpyxl.worksheet.worksheet.Worksheet,
    header_row: int,
    code_row: int,
) -> tuple[list[tuple[int, str]], int]:
    total_column = None
    subjects: list[tuple[int, str]] = []

    for column in range(4, ws.max_column + 1):
        subject_name = ws.cell(row=header_row, column=column).value
        if subject_name == "TOTAL %":
            total_column = column
            break
        if subject_name in {None, ""}:
            continue
        subject_code = ws.cell(row=code_row, column=column).value
        code = "" if subject_code in {None, "-"} else str(subject_code).strip()
        label = f"{code} {str(subject_name).strip()}".strip()
        subjects.append((column, label))

    if total_column is None:
        raise ValueError(f"Could not find TOTAL % column in sheet {ws.title}.")

    return subjects, total_column


def _attendance_value(value: object) -> float | None:
    if value in {None, "-", ""}:
        return None
    text = str(value).strip()
    if "(" in text and "%" in text:
        start = text.index("(") + 1
        end = text.index("%", start)
        return round(float(text[start:end]), 2)
    numeric = float(value)
    return round(numeric * 100, 2) if numeric <= 1 else round(numeric, 2)


def _section_from_sheet(title: str) -> str:
    match = re.search(r"SEC-(\d+)", title, re.IGNORECASE)
    if not match:
        raise ValueError(f"Could not determine section from sheet name {title}.")
    return match.group(1)


def import_multi_section_attendance_workbook(
    source_path: Path = SOURCE_PATH,
    csv_source_path: Path = CSV_SOURCE_PATH,
) -> None:
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    csv_lookup = _csv_lookup(csv_source_path)

    db = SessionLocal()
    try:
        inserted = 0
        updated = 0
        attendance_updates = 0

        for sheet_name in workbook.sheetnames:
            ws = workbook[sheet_name]
            section = _section_from_sheet(sheet_name)
            header_row, code_row, conduct_row = _header_rows(ws)
            subjects, total_column = _subject_columns(ws, header_row, code_row)
            data_start_row = conduct_row + 1

            for row_number in range(data_start_row, ws.max_row + 1):
                registration_number = ws.cell(row=row_number, column=2).value
                student_name = ws.cell(row=row_number, column=3).value
                overall_attendance = ws.cell(row=row_number, column=total_column).value

                if not registration_number or not student_name:
                    continue

                registration_number = str(registration_number).strip()
                csv_row = csv_lookup.get(registration_number, {})

                student = db.query(Student).filter(Student.registration_number == registration_number).first()
                if student is None:
                    student = Student(
                        registration_number=registration_number,
                        name=str(student_name).strip().title(),
                        email=_normalize_email(csv_row.get("email", ""), registration_number),
                        counselor_name=(csv_row.get("counsellor_name") or "-").strip() or "-",
                        codechef_username="-",
                        codechef_contests_participated=0,
                        codechef_problems_solved=0,
                        codechef_participation_status="Not Available",
                        section=(csv_row.get("sectioncode") or section).strip() or section,
                        gender=(csv_row.get("gender") or "-").strip().upper() or "-",
                        age=int(float(csv_row["age"])) if (csv_row.get("age") or "").strip() not in {"", "nan"} else None,
                        gpa=float(csv_row["cgpa"]) if (csv_row.get("cgpa") or "").strip() else 0.0,
                        attendance=float(overall_attendance or 0.0),
                        marks=0.0,
                        department="CSE",
                        year=int(float(csv_row["year"])) if (csv_row.get("year") or "").strip() else 3,
                        career_interest="-",
                        skills="-",
                    )
                    student.lms_activity = LMSActivity(
                        weekly_logins=0,
                        avg_time_spent=0.0,
                        assignment_submission_rate=0.0,
                        missed_assignments=0,
                    )
                    student.financial = Financial(
                        fee_due=0.0,
                        payment_delay_days=0,
                        scholarship_amount=0.0,
                    )
                    db.add(student)
                    db.flush()
                    inserted += 1
                else:
                    student.name = str(student_name).strip().title()
                    student.email = _normalize_email(csv_row.get("email", student.email), registration_number)
                    student.counselor_name = (csv_row.get("counsellor_name") or student.counselor_name or "-").strip() or "-"
                    student.section = (csv_row.get("sectioncode") or section or student.section).strip() or section
                    student.gender = (csv_row.get("gender") or student.gender or "-").strip().upper() or "-"
                    student.age = int(float(csv_row["age"])) if (csv_row.get("age") or "").strip() not in {"", "nan"} else student.age
                    if (csv_row.get("cgpa") or "").strip():
                        student.gpa = float(csv_row["cgpa"])
                    if (csv_row.get("year") or "").strip():
                        student.year = int(float(csv_row["year"]))
                    student.department = "CSE"
                    student.attendance = float(overall_attendance or 0.0)
                    updated += 1

                db.query(SubjectAttendance).filter(SubjectAttendance.student_id == student.id).delete(
                    synchronize_session=False
                )

                for column, subject_label in subjects:
                    attendance_percentage = _attendance_value(ws.cell(row=row_number, column=column).value)
                    if attendance_percentage is None:
                        continue
                    db.add(
                        SubjectAttendance(
                            student_id=student.id,
                            subject_name=subject_label,
                            attendance_percentage=attendance_percentage,
                        )
                    )

                attendance_updates += 1

        db.commit()
        print(
            f"Imported {inserted} students and updated {updated} students from {source_path.name}."
        )
        print(f"Rebuilt subject-wise attendance for {attendance_updates} student rows across {len(workbook.sheetnames)} sheets.")
    finally:
        db.close()


if __name__ == "__main__":
    cli_source_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else SOURCE_PATH
    import_multi_section_attendance_workbook(cli_source_path)
