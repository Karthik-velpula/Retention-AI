import random

from openpyxl import load_workbook

from app.db.session import SessionLocal
from app.models.student import Student
from app.models.subject_attendance import SubjectAttendance
from app.utils.init_db import init_db

TOTAL_REVISED_MARKS = 60
OPEN_ELECTIVE_ALLOCATION_PATH = "/Users/karthi/Downloads/III yr Open Electives students list.xlsx"
COMMON_INTERNAL_MARK_SUBJECT_CODES = {
    "22CS307",
    "22CS311",
    "22CS407",
}
T5_ONLY_SUBJECT_CODES = {
    
}
T5_ONLY_SUBJECT_LABELS = {
    
}
DEPARTMENT_ELECTIVE_SUBJECT_CODES = {
    "22CS801",
    "22CS802",
    "22CS804",
    "22CS805",
    "22CS806",
    "22CS808",
}
OPEN_ELECTIVE_SUBJECT_CODES = {
    "22BM854",
    "22BI851",
    "22CE852",
    "22CE855",
    "22CT851",
    "22CT853",
    "22EC856",
    "22EE855",
    "22FT853",
    "22MS851",
    "22MT855",
    "22MT861",
    "22PY857",
    "22RH301",
    "22TP851",
    "22TP852",
    "22TT854",
    "24RA852",
    "24RA853",
}
OPEN_ELECTIVE_CODE_ALIASES = {
    "24RA852": "24RA853",
}
OPEN_ELECTIVE_SUBJECT_LABELS = {
    "22BI851": "22BI851 EBI",
}
OPEN_ELECTIVE_MANUAL_OVERRIDES = {
    "231FA04283": "22BI851",
    "231FA04D22": "22BM854",
    "231FA04F10": "22BI851",
    "231FA04F94": "22BI851",
    "231FA04G70": "22BI851",
    "231FA04G82": "22BI851",
}
RAW_MARK_LIMITS = {
    "pre_t1_marks": 10,
    "t1_marks": 20,
    "t2_marks": 5,
    "t3_marks": 5,
    "t4_marks": 20,
}
T5_ASSIGNMENT_COUNT = 4
T5_ASSIGNMENT_MAX = 20


def _random_score(max_marks: int) -> float:
    lower_bound = max(0, round(max_marks * 0.55))
    return float(random.randint(lower_bound, max_marks))


def _random_t5_average() -> float:
    assignment_scores = [_random_score(T5_ASSIGNMENT_MAX) for _ in range(T5_ASSIGNMENT_COUNT)]
    return round(sum(assignment_scores) / T5_ASSIGNMENT_COUNT, 2)


def _random_t5_assignment_scores() -> list[float]:
    return [_random_score(T5_ASSIGNMENT_MAX) for _ in range(T5_ASSIGNMENT_COUNT)]


def _scaled_total_from_raw_scores(
    pre_t1_marks: float,
    t1_marks: float,
    t2_marks: float,
    t3_marks: float,
    t4_marks: float,
    t5_marks: float,
) -> float:
    return round(
        (pre_t1_marks / 10) * 5
        + (t1_marks / 20) * 8
        + (t2_marks / 5) * 3
        + (t3_marks / 5) * 3
        + t4_marks
        + t5_marks,
        2,
    )


def _subject_code(subject_name: str) -> str:
    return (subject_name or "").strip().split()[0].upper()


def _load_open_elective_allocations() -> dict[str, str]:
    workbook = load_workbook(OPEN_ELECTIVE_ALLOCATION_PATH, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]
    allocations: dict[str, str] = {}
    for row in worksheet.iter_rows(min_row=6, values_only=True):
        if not row or not row[1] or not row[2]:
            continue
        registration_number = str(row[1]).strip()
        raw_value = str(row[2]).strip()
        if raw_value.lower() == "not registered":
            continue
        if "(" not in raw_value or ")" not in raw_value:
            continue
        inside = raw_value.split("(", 1)[1].split(")", 1)[0]
        code = inside.replace(",", "").strip().upper()
        if not code:
            continue
        allocations[registration_number] = OPEN_ELECTIVE_CODE_ALIASES.get(code, code)
    allocations.update(OPEN_ELECTIVE_MANUAL_OVERRIDES)
    return allocations


def _is_common_subject(subject_name: str) -> bool:
    return _subject_code(subject_name) in COMMON_INTERNAL_MARK_SUBJECT_CODES


def _is_department_elective(subject_name: str) -> bool:
    return _subject_code(subject_name) in DEPARTMENT_ELECTIVE_SUBJECT_CODES


def _is_open_elective(subject_name: str) -> bool:
    return _subject_code(subject_name) in OPEN_ELECTIVE_SUBJECT_CODES


def _is_t5_only_subject(subject_name: str) -> bool:
    return _subject_code(subject_name) in T5_ONLY_SUBJECT_CODES


def _ensure_t5_only_subject_rows(
    db,
    student_id: int,
    subject_rows: list[SubjectAttendance],
) -> list[SubjectAttendance]:
    existing_codes = {_subject_code(subject_row.subject_name) for subject_row in subject_rows}
    updated_rows = list(subject_rows)

    for subject_code, subject_label in T5_ONLY_SUBJECT_LABELS.items():
        if subject_code in existing_codes:
            continue
        subject_row = SubjectAttendance(
            student_id=student_id,
            subject_name=subject_label,
            attendance_percentage=0,
        )
        db.add(subject_row)
        db.flush()
        updated_rows.append(subject_row)
        existing_codes.add(subject_code)

    return updated_rows


def _ensure_open_elective_subject_row(
    db,
    student_id: int,
    subject_rows: list[SubjectAttendance],
    allocated_open_elective_code: str | None,
) -> list[SubjectAttendance]:
    if not allocated_open_elective_code:
        return subject_rows

    existing_codes = {_subject_code(subject_row.subject_name) for subject_row in subject_rows}
    if allocated_open_elective_code in existing_codes:
        return subject_rows

    subject_label = OPEN_ELECTIVE_SUBJECT_LABELS.get(allocated_open_elective_code, allocated_open_elective_code)
    subject_row = SubjectAttendance(
        student_id=student_id,
        subject_name=subject_label,
        attendance_percentage=0,
    )
    db.add(subject_row)
    db.flush()
    return [*subject_rows, subject_row]


def populate_random_assessment_marks() -> dict[str, int]:
    init_db()
    db = SessionLocal()
    try:
        open_elective_allocations = _load_open_elective_allocations()
        students = db.query(Student).all()
        subject_rows_updated = 0

        for student in students:
            subject_rows = (
                db.query(SubjectAttendance)
                .filter(SubjectAttendance.student_id == student.id)
                .all()
            )

            if not subject_rows:
                student.pre_t1_marks = 0
                student.t1_marks = 0
                student.t2_marks = 0
                student.t3_marks = 0
                student.t4_marks = 0
                student.t5_marks = 0
                student.marks = 0
                continue

            pre_t1_total = 0.0
            t1_total = 0.0
            t2_total = 0.0
            t3_total = 0.0
            t4_total = 0.0
            t5_total = 0.0
            scaled_total = 0.0

            department_elective_assigned = False
            open_elective_assigned = False
            allocated_open_elective_code = open_elective_allocations.get(student.registration_number)
            subject_rows = _ensure_t5_only_subject_rows(
                db,
                student.id,
                subject_rows,
            )
            subject_rows = _ensure_open_elective_subject_row(
                db,
                student.id,
                subject_rows,
                allocated_open_elective_code,
            )

            for subject_row in sorted(subject_rows, key=lambda row: row.subject_name):
                include_subject = False
                subject_code = _subject_code(subject_row.subject_name)
                if _is_common_subject(subject_row.subject_name):
                    include_subject = True
                elif _is_t5_only_subject(subject_row.subject_name):
                    include_subject = True
                elif _is_department_elective(subject_row.subject_name) and not department_elective_assigned:
                    include_subject = True
                    department_elective_assigned = True
                elif (
                    _is_open_elective(subject_row.subject_name)
                    and not open_elective_assigned
                    and allocated_open_elective_code
                    and subject_code == allocated_open_elective_code
                ):
                    include_subject = True
                    open_elective_assigned = True

                if not include_subject:
                    subject_row.pre_t1_marks = 0
                    subject_row.t1_marks = 0
                    subject_row.t2_marks = 0
                    subject_row.t3_marks = 0
                    subject_row.t4_marks = 0
                    subject_row.t5_marks = 0
                    subject_row.t5_assignment_1 = 0
                    subject_row.t5_assignment_2 = 0
                    subject_row.t5_assignment_3 = 0
                    subject_row.t5_assignment_4 = 0
                    subject_row.total_marks = 0
                    continue

                if _is_t5_only_subject(subject_row.subject_name):
                    subject_row.pre_t1_marks = 0
                    subject_row.t1_marks = 0
                    subject_row.t2_marks = 0
                    subject_row.t3_marks = 0
                    subject_row.t4_marks = 0
                else:
                    for field_name, max_marks in RAW_MARK_LIMITS.items():
                        setattr(subject_row, field_name, _random_score(max_marks))
                t5_assignment_scores = _random_t5_assignment_scores()
                (
                    subject_row.t5_assignment_1,
                    subject_row.t5_assignment_2,
                    subject_row.t5_assignment_3,
                    subject_row.t5_assignment_4,
                ) = t5_assignment_scores
                subject_row.t5_marks = round(sum(t5_assignment_scores) / T5_ASSIGNMENT_COUNT, 2)
                if _is_t5_only_subject(subject_row.subject_name):
                    subject_row.total_marks = subject_row.t5_marks
                else:
                    subject_row.total_marks = _scaled_total_from_raw_scores(
                        subject_row.pre_t1_marks,
                        subject_row.t1_marks,
                        subject_row.t2_marks,
                        subject_row.t3_marks,
                        subject_row.t4_marks,
                        subject_row.t5_marks,
                    )

                pre_t1_total += subject_row.pre_t1_marks
                t1_total += subject_row.t1_marks
                t2_total += subject_row.t2_marks
                t3_total += subject_row.t3_marks
                t4_total += subject_row.t4_marks
                t5_total += subject_row.t5_marks
                scaled_total += subject_row.total_marks
                subject_rows_updated += 1

            selected_subject_count = sum(1 for subject_row in subject_rows if subject_row.total_marks > 0)
            subject_count = max(1, selected_subject_count)
            student.pre_t1_marks = round(pre_t1_total / subject_count, 2)
            student.t1_marks = round(t1_total / subject_count, 2)
            student.t2_marks = round(t2_total / subject_count, 2)
            student.t3_marks = round(t3_total / subject_count, 2)
            student.t4_marks = round(t4_total / subject_count, 2)
            student.t5_marks = round(t5_total / subject_count, 2)
            average_scaled_total = round(scaled_total / subject_count, 2)
            student.marks = round((average_scaled_total / TOTAL_REVISED_MARKS) * 100, 2)

        db.commit()
        return {"updated_students": len(students), "updated_subject_rows": subject_rows_updated}
    finally:
        db.close()


if __name__ == "__main__":
    summary = populate_random_assessment_marks()
    print(
        f"Updated random assessment marks for {summary['updated_students']} students across "
        f"{summary['updated_subject_rows']} subject rows."
    )
