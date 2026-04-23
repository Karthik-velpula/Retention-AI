from __future__ import annotations

import csv
import sys
from pathlib import Path

import openpyxl
from email_validator import EmailNotValidError, validate_email

from app.db.session import SessionLocal
from app.models.financial import Financial
from app.models.lms_activity import LMSActivity
from app.models.student import Student
from app.models.subject_attendance import SubjectAttendance

SOURCE_PATH = Path("/Users/karthi/Downloads/Subjectwiseattendence(02-04-26).xlsx")
CSV_SOURCE_PATH = Path("/Users/karthi/Downloads/studentdetails1.csv")


def _normalize_email(raw_email: str, registration_number: str) -> str:
    value = str(raw_email).strip().lower().replace(" ", "")
    if not value or value in {"-", "0", "nan"}:
        return f"{registration_number.lower()}@gmail.com"
    try:
        return validate_email(value, check_deliverability=False).normalized
    except EmailNotValidError:
        return f"{registration_number.lower()}@gmail.com"


def _csv_lookup(source_path: Path) -> dict[str, dict[str, str]]:
    lookup: dict[str, dict[str, str]] = {}
    with source_path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        for row in reader:
            registration_number = (row.get("registerno") or "").strip()
            if registration_number:
                lookup[registration_number] = row
    return lookup


def _subject_columns(ws: openpyxl.worksheet.worksheet.Worksheet) -> list[tuple[int, str]]:
    subjects: list[tuple[int, str]] = []
    for column in range(4, ws.max_column):
        subject_name = ws.cell(row=5, column=column).value
        subject_code = ws.cell(row=6, column=column).value
        if subject_name in {None, "TOTAL %"}:
            continue
        label = f"{str(subject_code).strip()} {str(subject_name).strip()}".strip()
        subjects.append((column, label))
    return subjects


def import_subjectwise_xlsx(
    source_path: Path = SOURCE_PATH,
    csv_source_path: Path = CSV_SOURCE_PATH,
) -> None:
    workbook = openpyxl.load_workbook(source_path, data_only=True)
    worksheet = workbook[workbook.sheetnames[0]]

    csv_lookup = _csv_lookup(csv_source_path)
    subjects = _subject_columns(worksheet)

    db = SessionLocal()
    try:
        inserted = 0
        updated = 0

        for row_number in range(8, worksheet.max_row + 1):
            registration_number = worksheet.cell(row=row_number, column=2).value
            student_name = worksheet.cell(row=row_number, column=3).value
            overall_attendance = worksheet.cell(row=row_number, column=28).value

            if not registration_number or not student_name:
                continue

            registration_number = str(registration_number).strip()
            csv_row = csv_lookup.get(registration_number, {})

            student = db.query(Student).filter(Student.registration_number == registration_number).first()
            if student is None:
                student = Student(
                    registration_number=registration_number,
                    name=str(student_name).strip().title(),
                    email=_normalize_email(csv_row.get("email", ""), registration_number),
                    counselor_name=(csv_row.get("counsellor_name") or "-").strip() or "-",
                    codechef_username="-",
                    codechef_contests_participated=0,
                    codechef_problems_solved=0,
                    codechef_participation_status="Not Available",
                    section=(csv_row.get("sectioncode") or "12").strip() or "12",
                    gender=(csv_row.get("gender") or "-").strip().upper() or "-",
                    age=int(float(csv_row["age"])) if (csv_row.get("age") or "").strip() not in {"", "nan"} else None,
                    gpa=float(csv_row["cgpa"]) if (csv_row.get("cgpa") or "").strip() else 0.0,
                    attendance=float(overall_attendance or 0.0),
                    marks=0.0,
                    department="CSE",
                    year=int(float(csv_row["year"])) if (csv_row.get("year") or "").strip() else 3,
                    career_interest="-",
                    skills="-",
                )
                student.lms_activity = LMSActivity(
                    weekly_logins=0,
                    avg_time_spent=0.0,
                    assignment_submission_rate=0.0,
                    missed_assignments=0,
                )
                student.financial = Financial(
                    fee_due=0.0,
                    payment_delay_days=0,
                    scholarship_amount=0.0,
                )
                db.add(student)
                db.flush()
                inserted += 1
            else:
                student.name = str(student_name).strip().title()
                student.email = _normalize_email(csv_row.get("email", student.email), registration_number)
                student.counselor_name = (csv_row.get("counsellor_name") or student.counselor_name or "-").strip() or "-"
                student.section = (csv_row.get("sectioncode") or student.section or "12").strip() or "12"
                student.gender = (csv_row.get("gender") or student.gender or "-").strip().upper() or "-"
                student.age = int(float(csv_row["age"])) if (csv_row.get("age") or "").strip() not in {"", "nan"} else student.age
                if (csv_row.get("cgpa") or "").strip():
                    student.gpa = float(csv_row["cgpa"])
                if (csv_row.get("year") or "").strip():
                    student.year = int(float(csv_row["year"]))
                student.department = "CSE"
                student.attendance = float(overall_attendance or 0.0)
                updated += 1

            db.query(SubjectAttendance).filter(SubjectAttendance.student_id == student.id).delete(
                synchronize_session=False
            )

            for column, subject_label in subjects:
                value = worksheet.cell(row=row_number, column=column).value
                if value in {None, "-", ""}:
                    continue
                attendance_percentage = float(value) * 100 if float(value) <= 1 else float(value)
                db.add(
                    SubjectAttendance(
                        student_id=student.id,
                        subject_name=subject_label,
                        attendance_percentage=round(attendance_percentage, 2),
                    )
                )

        db.commit()
        print(
            f"Imported {inserted} students and updated {updated} students from {source_path.name}."
        )
        print(f"Stored subject-wise attendance using {len(subjects)} subject columns.")
    finally:
        db.close()


if __name__ == "__main__":
    cli_source_path = Path(sys.argv[1]).expanduser() if len(sys.argv) > 1 else SOURCE_PATH
    import_subjectwise_xlsx(cli_source_path)
